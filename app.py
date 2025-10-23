from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os
import zipfile
import tempfile
from datetime import datetime
from functools import wraps
import uuid
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sponsoring.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# Helper function for European number formatting
def format_european_currency(amount):
    """Format currency in European style: 1.234,56"""
    if amount is None:
        return "0,00"
    # Convert to string with 2 decimal places
    formatted = f"{amount:,.2f}"
    # Replace comma with dot for thousands separator, and dot with comma for decimal
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted

# Helper function for getting display amount with fallback
def get_display_amount(sponsoring):
    """Get the amount to display, using fallback logic.
    Add bedrag_kaarten and netto_bedrag_excl_btw together.
    If one is not filled in, use 0 for that amount.
    """
    netto = sponsoring.netto_bedrag_excl_btw or 0
    kaarten = sponsoring.bedrag_kaarten or 0
    
    return netto + kaarten

# Make the helper functions available in templates
app.jinja_env.globals.update(
    format_european_currency=format_european_currency,
    get_display_amount=get_display_amount
)

# Database Models
class Evenement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    evenementcode = db.Column(db.String(50), nullable=False, unique=True)
    naam = db.Column(db.String(200), nullable=False)
    datum = db.Column(db.Date, nullable=False)
    locatie = db.Column(db.String(200), nullable=False)
    omschrijving = db.Column(db.Text)
    
    # Relationships
    kontrakten = db.relationship('Kontrakt', backref='evenement', lazy=True, cascade='all, delete-orphan')
    sponsoringen = db.relationship('Sponsoring', backref='evenement', lazy=True, cascade='all, delete-orphan')

class Bestuurslid(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    initialen = db.Column(db.String(10), nullable=False)
    naam = db.Column(db.String(100), nullable=True)
    
    # Relationships
    sponsors = db.relationship('Sponsor', backref='bestuurslid', lazy=True)
    sponsoringen = db.relationship('Sponsoring', backref='aangebracht_door', lazy=True)

class Kontrakt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    evenement_id = db.Column(db.Integer, db.ForeignKey('evenement.id'), nullable=False)
    kontrakt = db.Column(db.String(100), nullable=False)
    bedrag = db.Column(db.Float, nullable=False)
    tegenprestatie = db.Column(db.Text)
    
    # Relationships
    sponsoringen = db.relationship('Sponsoring', backref='kontrakt', lazy=True, cascade='all, delete-orphan')
    
    # Unique constraint on evenement + kontrakt
    __table_args__ = (db.UniqueConstraint('evenement_id', 'kontrakt', name='unique_evenement_kontrakt'),)

class Sponsor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    naam = db.Column(db.String(200), nullable=False)
    straat = db.Column(db.String(100))
    huisnummer = db.Column(db.String(10))
    postcode = db.Column(db.String(10))
    gemeente = db.Column(db.String(100))
    kontaktpersoon = db.Column(db.String(100))
    telefoon = db.Column(db.String(20))
    email = db.Column(db.String(100))
    btw_nummer = db.Column(db.String(50))
    bestuurslid_id = db.Column(db.Integer, db.ForeignKey('bestuurslid.id'))
    opmerkingen = db.Column(db.Text)
    
    # Relationships
    sponsoringen = db.relationship('Sponsoring', backref='sponsor', lazy=True, cascade='all, delete-orphan')

class Gebruiker(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=True)  # Nullable for new users
    eerste_aanmelding = db.Column(db.DateTime, default=datetime.utcnow)
    laatste_activiteit = db.Column(db.DateTime, default=datetime.utcnow)
    rol = db.Column(db.String(20), nullable=False, default='lezer')  # beheerder, gebruiker, lezer
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)
    
    def is_beheerder(self):
        return self.rol == 'beheerder'
    
    def is_gebruiker(self):
        return self.rol in ['beheerder', 'gebruiker']
    
    def is_lezer(self):
        return self.rol in ['beheerder', 'gebruiker', 'lezer']

class Sponsoring(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    evenement_id = db.Column(db.Integer, db.ForeignKey('evenement.id'), nullable=False)
    kontrakt_id = db.Column(db.Integer, db.ForeignKey('kontrakt.id'), nullable=False)
    sponsor_id = db.Column(db.Integer, db.ForeignKey('sponsor.id'), nullable=False)
    aangebracht_door_id = db.Column(db.Integer, db.ForeignKey('bestuurslid.id'), nullable=False)
    bedrag_kaarten = db.Column(db.Float)
    netto_bedrag_excl_btw = db.Column(db.Float)
    facturatiebedrag_incl_btw = db.Column(db.Float)
    gefactureerd = db.Column(db.Boolean, default=False)
    betaald = db.Column(db.Boolean, default=False)
    opmerkingen = db.Column(db.Text)
    logo_bezorgd = db.Column(db.Boolean, default=False)
    logo_afgewerkt = db.Column(db.Boolean, default=False)
    logo_origineel = db.Column(db.String(255))
    logo_afgewerkt_file = db.Column(db.String(255))

# Context processor to make current_user available in all templates
@app.context_processor
def inject_user():
    if 'user_id' in session:
        user = Gebruiker.query.get(session['user_id'])
        return dict(current_user=user)
    return dict(current_user=None)

# Authentication decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def beheerder_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = Gebruiker.query.get(session['user_id'])
        if not user or not user.is_beheerder():
            flash('Je hebt geen toegang tot deze functie.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def gebruiker_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = Gebruiker.query.get(session['user_id'])
        if not user or not user.is_gebruiker():
            flash('Je hebt geen toegang tot deze functie.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = Gebruiker.query.filter_by(email=email).first()
        
        if user:
            # Check if user has no password set (new user)
            if not user.password_hash:
                # Allow login without password for new users (password can be empty or anything)
                session['user_id'] = user.id
                session['user_email'] = user.email
                user.laatste_activiteit = datetime.utcnow()
                db.session.commit()
                flash('Welkom! Je moet eerst je wachtwoord instellen.', 'info')
                return redirect(url_for('set_password'))
            elif password and user.check_password(password):
                # Normal login with password
                session['user_id'] = user.id
                session['user_email'] = user.email
                user.laatste_activiteit = datetime.utcnow()
                db.session.commit()
                flash('Succesvol ingelogd!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Ongeldige email of wachtwoord.', 'error')
        else:
            flash('Ongeldige email of wachtwoord.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('user_email', None)
    flash('Je bent uitgelogd.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    user = Gebruiker.query.get(session['user_id'])
    
    # Check if user needs to set password
    if not user.password_hash:
        return redirect(url_for('set_password'))
    
    # Get dashboard statistics
    evenementen_count = Evenement.query.count()
    sponsoringen_count = Sponsoring.query.count()
    sponsors_count = Sponsor.query.count()
    bestuursleden_count = Bestuurslid.query.count()
    
    return render_template('dashboard.html', 
                         evenementen_count=evenementen_count,
                         sponsoringen_count=sponsoringen_count,
                         sponsors_count=sponsors_count,
                         bestuursleden_count=bestuursleden_count)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        user = Gebruiker.query.get(session['user_id'])
        
        if not user.check_password(current_password):
            flash('Huidig wachtwoord is onjuist.', 'error')
            return render_template('change_password.html')
        
        if new_password != confirm_password:
            flash('Nieuwe wachtwoorden komen niet overeen.', 'error')
            return render_template('change_password.html')
        
        if len(new_password) < 6:
            flash('Wachtwoord moet minimaal 6 karakters lang zijn.', 'error')
            return render_template('change_password.html')
        
        user.set_password(new_password)
        db.session.commit()
        flash('Wachtwoord succesvol gewijzigd!', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('change_password.html')

@app.route('/set-password', methods=['GET', 'POST'])
@login_required
def set_password():
    user = Gebruiker.query.get(session['user_id'])
    
    if user.password_hash:  # User already has a password
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        
        if password != confirm_password:
            flash('Wachtwoorden komen niet overeen.', 'error')
            return render_template('set_password.html')
        
        if len(password) < 6:
            flash('Wachtwoord moet minimaal 6 karakters lang zijn.', 'error')
            return render_template('set_password.html')
        
        user.set_password(password)
        db.session.commit()
        flash('Wachtwoord succesvol ingesteld!', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('set_password.html')

# User management routes (admin only)
@app.route('/gebruikers')
@beheerder_required
def gebruikers():
    gebruikers = Gebruiker.query.order_by(Gebruiker.email).all()
    return render_template('gebruikers.html', gebruikers=gebruikers)

@app.route('/gebruikers/toevoegen', methods=['GET', 'POST'])
@beheerder_required
def gebruiker_toevoegen():
    if request.method == 'POST':
        email = request.form['email']
        rol = request.form['rol']
        
        # Check if user already exists
        if Gebruiker.query.filter_by(email=email).first():
            flash('Een gebruiker met dit email adres bestaat al.', 'error')
            return render_template('gebruiker_toevoegen.html')
        
        # Create new user
        gebruiker = Gebruiker(
            email=email,
            rol=rol
            # password_hash remains None for first-time setup
        )
        db.session.add(gebruiker)
        db.session.commit()
        
        flash(f'Gebruiker {email} is toegevoegd. Ze moeten hun wachtwoord instellen bij eerste login.', 'success')
        return redirect(url_for('gebruikers'))
    
    return render_template('gebruiker_toevoegen.html')

@app.route('/gebruikers/<int:user_id>/bewerken', methods=['GET', 'POST'])
@beheerder_required
def gebruiker_bewerken(user_id):
    gebruiker = Gebruiker.query.get_or_404(user_id)
    
    if request.method == 'POST':
        gebruiker.email = request.form['email']
        gebruiker.rol = request.form['rol']
        db.session.commit()
        
        flash('Gebruiker succesvol bijgewerkt.', 'success')
        return redirect(url_for('gebruikers'))
    
    return render_template('gebruiker_bewerken.html', gebruiker=gebruiker)

@app.route('/gebruikers/<int:user_id>/verwijderen', methods=['POST'])
@beheerder_required
def gebruiker_verwijderen(user_id):
    gebruiker = Gebruiker.query.get_or_404(user_id)
    
    # Prevent deleting yourself
    if gebruiker.id == session['user_id']:
        flash('Je kunt jezelf niet verwijderen.', 'error')
        return redirect(url_for('gebruikers'))
    
    # Prevent deleting the last admin
    if gebruiker.rol == 'beheerder' and Gebruiker.query.filter_by(rol='beheerder').count() <= 1:
        flash('Je kunt de laatste beheerder niet verwijderen.', 'error')
        return redirect(url_for('gebruikers'))
    
    db.session.delete(gebruiker)
    db.session.commit()
    
    flash('Gebruiker succesvol verwijderd.', 'success')
    return redirect(url_for('gebruikers'))

@app.route('/gebruikers/<int:user_id>/wachtwoord-reset', methods=['POST'])
@beheerder_required
def gebruiker_wachtwoord_reset(user_id):
    gebruiker = Gebruiker.query.get_or_404(user_id)
    
    # Reset password by setting password_hash to None
    gebruiker.password_hash = None
    db.session.commit()
    
    flash(f'Wachtwoord van {gebruiker.email} is gereset. De gebruiker moet een nieuw wachtwoord instellen bij de volgende login.', 'success')
    return redirect(url_for('gebruikers'))

# Routes
@app.route('/')
def index():
    return redirect(url_for('dashboard'))
    
    # Get recent data
    recente_evenementen = Evenement.query.order_by(Evenement.datum.desc()).limit(5).all()
    recente_sponsoringen = Sponsoring.query.join(Evenement).order_by(Evenement.datum.desc()).limit(5).all()
    
    return render_template('index.html',
                         evenementen_count=evenementen_count,
                         sponsoringen_count=sponsoringen_count,
                         sponsors_count=sponsors_count,
                         bestuursleden_count=bestuursleden_count,
                         recente_evenementen=recente_evenementen,
                         recente_sponsoringen=recente_sponsoringen)

@app.route('/statistieken')
@login_required
def statistieken():
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    
    # Calculate statistics for each evenement
    evenement_stats = []
    for evenement in evenementen:
        kontrakt_stats = []
        total_excl_btw = 0
        total_incl_btw = 0
        total_betaald = 0
        total_openstaand = 0
        
        for kontrakt in evenement.kontrakten:
            # Calculate totals for this kontrakt
            kontrakt_excl_btw = 0
            kontrakt_incl_btw = 0
            kontrakt_betaald = 0
            kontrakt_openstaand = 0
            
            for sponsoring in kontrakt.sponsoringen:
                # Use fallback logic for excl BTW
                display_amount = get_display_amount(sponsoring)
                kontrakt_excl_btw += display_amount
                
                if sponsoring.facturatiebedrag_incl_btw:
                    kontrakt_incl_btw += sponsoring.facturatiebedrag_incl_btw
                    
                    # Calculate betaald and openstaand
                    if sponsoring.betaald:
                        kontrakt_betaald += sponsoring.facturatiebedrag_incl_btw
                    else:
                        kontrakt_openstaand += sponsoring.facturatiebedrag_incl_btw
            
            if kontrakt_excl_btw > 0 or kontrakt_incl_btw > 0:
                kontrakt_stats.append({
                    'kontrakt': kontrakt.kontrakt,
                    'excl_btw': kontrakt_excl_btw,
                    'incl_btw': kontrakt_incl_btw,
                    'betaald': kontrakt_betaald,
                    'openstaand': kontrakt_openstaand
                })
                total_excl_btw += kontrakt_excl_btw
                total_incl_btw += kontrakt_incl_btw
                total_betaald += kontrakt_betaald
                total_openstaand += kontrakt_openstaand
        
        evenement_stats.append({
            'evenement': evenement,
            'kontrakt_stats': kontrakt_stats,
            'total_excl_btw': total_excl_btw,
            'total_incl_btw': total_incl_btw,
            'total_betaald': total_betaald,
            'total_openstaand': total_openstaand
        })
    
    return render_template('statistieken.html', evenement_stats=evenement_stats)

@app.route('/evenementen')
@login_required
def evenementen():
    sort_key = request.args.get('sort', '')
    sort_dir = request.args.get('dir', 'asc')
    
    evenementen_list = Evenement.query.all()
    
    # Apply sorting
    def sort_key_fn(e):
        if sort_key == 'code':
            return (e.evenementcode or '').lower()
        if sort_key == 'naam':
            return (e.naam or '').lower()
        if sort_key == 'datum':
            return e.datum
        if sort_key == 'locatie':
            return (e.locatie or '').lower()
        if sort_key == 'kontrakten':
            return len(e.kontrakten)
        if sort_key == 'sponsoringen':
            return len(e.sponsoringen)
        # default: newest first
        return e.datum
    
    if sort_key:
        evenementen_list = sorted(evenementen_list, key=sort_key_fn, reverse=(sort_dir == 'desc'))
    else:
        evenementen_list = sorted(evenementen_list, key=lambda e: e.datum, reverse=True)
    
    return render_template('evenementen.html', 
                         evenementen=evenementen_list,
                         selected_sort=sort_key,
                         selected_dir=sort_dir)

@app.route('/evenementen/add', methods=['GET', 'POST'])
@gebruiker_required
def add_evenement():
    if request.method == 'POST':
        evenement = Evenement(
            evenementcode=request.form['evenementcode'],
            naam=request.form['naam'],
            datum=datetime.strptime(request.form['datum'], '%Y-%m-%d').date(),
            locatie=request.form['locatie'],
            omschrijving=request.form['omschrijving']
        )
        db.session.add(evenement)
        db.session.commit()
        flash('Evenement succesvol toegevoegd!', 'success')
        return redirect(url_for('evenementen'))
    return render_template('add_evenement.html')

@app.route('/evenementen/<int:id>')
@login_required
def evenement_detail(id):
    evenement = Evenement.query.get_or_404(id)
    
    # Calculate totals safely
    total_bedrag_incl_btw = sum(s.facturatiebedrag_incl_btw for s in evenement.sponsoringen if s.facturatiebedrag_incl_btw is not None)
    total_bedrag_excl_btw = sum(s.netto_bedrag_excl_btw for s in evenement.sponsoringen if s.netto_bedrag_excl_btw is not None)
    
    return render_template('evenement_detail.html', 
                         evenement=evenement,
                         total_bedrag_incl_btw=total_bedrag_incl_btw,
                         total_bedrag_excl_btw=total_bedrag_excl_btw)

@app.route('/evenementen/<int:id>/edit', methods=['GET', 'POST'])
@gebruiker_required
def edit_evenement(id):
    evenement = Evenement.query.get_or_404(id)
    if request.method == 'POST':
        evenement.evenementcode = request.form['evenementcode']
        evenement.naam = request.form['naam']
        evenement.datum = datetime.strptime(request.form['datum'], '%Y-%m-%d').date()
        evenement.locatie = request.form['locatie']
        evenement.omschrijving = request.form['omschrijving']
        db.session.commit()
        flash('Evenement succesvol bijgewerkt!', 'success')
        return redirect(url_for('evenement_detail', id=evenement.id))
    return render_template('edit_evenement.html', evenement=evenement)

@app.route('/evenementen/<int:id>/delete', methods=['POST'])
@beheerder_required
def delete_evenement(id):
    evenement = Evenement.query.get_or_404(id)
    
    # Delete all related sponsoringen first
    for sponsoring in evenement.sponsoringen:
        # Delete logo files if they exist
        if sponsoring.logo_origineel:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], sponsoring.logo_origineel))
            except:
                pass
        if sponsoring.logo_afgewerkt_file:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], sponsoring.logo_afgewerkt_file))
            except:
                pass
        db.session.delete(sponsoring)
    
    # Delete all related kontrakten
    for kontrakt in evenement.kontrakten:
        db.session.delete(kontrakt)
    
    # Delete the evenement itself
    db.session.delete(evenement)
    db.session.commit()
    
    flash(f'Evenement "{evenement.naam}" en alle gerelateerde data zijn succesvol verwijderd.', 'success')
    return redirect(url_for('evenementen'))

@app.route('/evenementen/export/excel')
@login_required
def export_evenementen_excel():
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Evenementen"
    
    # Headers
    headers = ['Evenementcode', 'Naam', 'Datum', 'Locatie', 'Aantal Kontrakten', 'Aantal Sponsoringen']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, evenement in enumerate(evenementen, 2):
        ws.cell(row=row, column=1, value=evenement.evenementcode)
        ws.cell(row=row, column=2, value=evenement.naam)
        ws.cell(row=row, column=3, value=evenement.datum.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=evenement.locatie)
        ws.cell(row=row, column=5, value=len(evenement.kontrakten))
        ws.cell(row=row, column=6, value=len(evenement.sponsoringen))
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='evenementen.xlsx')

@app.route('/evenementen/export/pdf')
@login_required
def export_evenementen_pdf():
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30)
    story.append(Paragraph("Evenementen Overzicht", title_style))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Evenementcode', 'Naam', 'Datum', 'Locatie', 'Kontrakten', 'Sponsoringen']]
    for evenement in evenementen:
        data.append([
            evenement.evenementcode,
            evenement.naam,
            evenement.datum.strftime('%d/%m/%Y'),
            evenement.locatie,
            str(len(evenement.kontrakten)),
            str(len(evenement.sponsoringen))
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='evenementen.pdf')

@app.route('/bestuursleden')
@login_required
def bestuursleden():
    sort_key = request.args.get('sort', '')
    sort_dir = request.args.get('dir', 'asc')
    
    bestuursleden = Bestuurslid.query.all()
    
    # Apply sorting
    def sort_key_fn(b):
        if sort_key == 'naam':
            return (b.naam or b.initialen or '').lower()
        if sort_key == 'initialen':
            return (b.initialen or '').lower()
        if sort_key == 'sponsors':
            return len(b.sponsors)
        if sort_key == 'sponsoringen':
            return len(b.sponsoringen)
        # default: alphabetical by name or initialen
        return (b.naam or b.initialen or '').lower()
    
    if sort_key:
        bestuursleden = sorted(bestuursleden, key=sort_key_fn, reverse=(sort_dir == 'desc'))
    else:
        bestuursleden = sorted(bestuursleden, key=lambda b: (b.naam or b.initialen or '').lower())
    
    return render_template('bestuursleden.html', 
                         bestuursleden=bestuursleden,
                         selected_sort=sort_key,
                         selected_dir=sort_dir)

@app.route('/bestuursleden/add', methods=['GET', 'POST'])
@gebruiker_required
def add_bestuurslid():
    if request.method == 'POST':
        bestuurslid = Bestuurslid(
            initialen=request.form['initialen'],
            naam=request.form['naam']
        )
        db.session.add(bestuurslid)
        db.session.commit()
        flash('Bestuurslid succesvol toegevoegd!', 'success')
        return redirect(url_for('bestuursleden'))
    return render_template('add_bestuurslid.html')

@app.route('/bestuursleden/<int:id>')
@login_required
def bestuurslid_detail(id):
    bestuurslid = Bestuurslid.query.get_or_404(id)
    # Calculate totals safely to avoid None issues in templates
    total_opgehaald = sum(
        s.facturatiebedrag_incl_btw for s in bestuurslid.sponsoringen if s.facturatiebedrag_incl_btw is not None
    )
    return render_template('bestuurslid_detail.html', bestuurslid=bestuurslid, total_opgehaald=total_opgehaald)

@app.route('/bestuursleden/<int:id>/edit', methods=['GET', 'POST'])
@gebruiker_required
def edit_bestuurslid(id):
    bestuurslid = Bestuurslid.query.get_or_404(id)
    if request.method == 'POST':
        bestuurslid.initialen = request.form['initialen']
        bestuurslid.naam = request.form['naam'] if request.form['naam'] else None
        db.session.commit()
        flash('Bestuurslid succesvol bijgewerkt!', 'success')
        return redirect(url_for('bestuurslid_detail', id=bestuurslid.id))
    return render_template('edit_bestuurslid.html', bestuurslid=bestuurslid)

@app.route('/bestuursleden/<int:id>/delete', methods=['POST'])
@beheerder_required
def delete_bestuurslid(id):
    bestuurslid = Bestuurslid.query.get_or_404(id)
    
    # Check if bestuurslid has related sponsoringen
    if bestuurslid.sponsoringen:
        flash(f'Bestuurslid "{bestuurslid.initialen}" kan niet worden verwijderd omdat er nog sponsoringen aan gekoppeld zijn.', 'error')
        return redirect(url_for('bestuursleden'))
    
    # Check if bestuurslid has related sponsors
    if bestuurslid.sponsors:
        flash(f'Bestuurslid "{bestuurslid.initialen}" kan niet worden verwijderd omdat er nog sponsors aan gekoppeld zijn.', 'error')
        return redirect(url_for('bestuursleden'))
    
    # Delete the bestuurslid
    db.session.delete(bestuurslid)
    db.session.commit()
    
    flash(f'Bestuurslid "{bestuurslid.initialen}" is succesvol verwijderd.', 'success')
    return redirect(url_for('bestuursleden'))

@app.route('/bestuursleden/export/excel')
@login_required
def export_bestuursleden_excel():
    bestuursleden = Bestuurslid.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bestuursleden"
    
    # Headers
    headers = ['Naam', 'Initialen', 'Aantal Sponsors', 'Aantal Sponsoringen']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, bestuurslid in enumerate(bestuursleden, 2):
        ws.cell(row=row, column=1, value=bestuurslid.naam or bestuurslid.initialen)
        ws.cell(row=row, column=2, value=bestuurslid.initialen)
        ws.cell(row=row, column=3, value=len(bestuurslid.sponsors))
        ws.cell(row=row, column=4, value=len(bestuurslid.sponsoringen))
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='bestuursleden.xlsx')

@app.route('/bestuursleden/export/pdf')
@login_required
def export_bestuursleden_pdf():
    bestuursleden = Bestuurslid.query.all()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30)
    story.append(Paragraph("Bestuursleden Overzicht", title_style))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Naam', 'Initialen', 'Sponsors', 'Sponsoringen']]
    for bestuurslid in bestuursleden:
        data.append([
            bestuurslid.naam or bestuurslid.initialen,
            bestuurslid.initialen,
            str(len(bestuurslid.sponsors)),
            str(len(bestuurslid.sponsoringen))
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='bestuursleden.pdf')

@app.route('/kontrakten')
@login_required
def kontrakten():
    # Get filter parameters
    evenement_filter = request.args.get('evenement', '')
    sort_key = request.args.get('sort', '')
    sort_dir = request.args.get('dir', 'asc')
    
    # If no evenement filter is provided, default to the most recent event
    if not evenement_filter:
        latest_evenement = Evenement.query.order_by(Evenement.datum.desc()).first()
        if latest_evenement:
            evenement_filter = str(latest_evenement.id)
    
    # Start with base query
    query = Kontrakt.query.join(Evenement)
    
    # Apply evenement filter if provided
    if evenement_filter:
        query = query.filter(Kontrakt.evenement_id == int(evenement_filter))
    
    kontrakten = query.all()
    
    # Apply sorting
    def sort_key_fn(k):
        if sort_key == 'evenement':
            return (k.evenement.naam or '').lower()
        if sort_key == 'kontrakt':
            return (k.kontrakt or '').lower()
        if sort_key == 'bedrag':
            return float(k.bedrag or 0)
        if sort_key == 'tegenprestatie':
            return (k.tegenprestatie or '').lower()
        if sort_key == 'sponsoringen':
            return len(k.sponsoringen)
        # default: newest first
        return k.evenement.datum
    
    if sort_key:
        kontrakten = sorted(kontrakten, key=sort_key_fn, reverse=(sort_dir == 'desc'))
    else:
        kontrakten = sorted(kontrakten, key=lambda k: k.evenement.datum, reverse=True)
    
    # Calculate total bedrag
    total_bedrag = sum(k.bedrag for k in kontrakten if k.bedrag is not None)
    
    # Get evenementen for filter dropdown
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    
    return render_template('kontrakten.html', 
                         kontrakten=kontrakten, 
                         total_bedrag=total_bedrag,
                         evenementen=evenementen,
                         selected_evenement=evenement_filter,
                         selected_sort=sort_key,
                         selected_dir=sort_dir)

@app.route('/kontrakten/add', methods=['GET', 'POST'])
@gebruiker_required
def add_kontrakt():
    if request.method == 'POST':
        kontrakt = Kontrakt(
            evenement_id=request.form['evenement_id'],
            kontrakt=request.form['kontrakt'],
            bedrag=float(request.form['bedrag']),
            tegenprestatie=request.form['tegenprestatie']
        )
        db.session.add(kontrakt)
        db.session.commit()
        flash('Kontrakt succesvol toegevoegd!', 'success')
        return redirect(url_for('kontrakten'))
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    return render_template('add_kontrakt.html', evenementen=evenementen)

@app.route('/kontrakten/<int:id>')
@login_required
def kontrakt_detail(id):
    kontrakt = Kontrakt.query.get_or_404(id)
    
    # Calculate total bedrag
    total_bedrag = sum(s.facturatiebedrag_incl_btw for s in kontrakt.sponsoringen if s.facturatiebedrag_incl_btw is not None)
    
    return render_template('kontrakt_detail.html', kontrakt=kontrakt, total_bedrag=total_bedrag)

@app.route('/kontrakten/<int:id>/edit', methods=['GET', 'POST'])
@gebruiker_required
def edit_kontrakt(id):
    kontrakt = Kontrakt.query.get_or_404(id)
    if request.method == 'POST':
        kontrakt.kontrakt = request.form['kontrakt']
        kontrakt.bedrag = float(request.form['bedrag'])
        kontrakt.tegenprestatie = request.form['tegenprestatie']
        db.session.commit()
        flash('Kontrakt succesvol bijgewerkt!', 'success')
        return redirect(url_for('kontrakt_detail', id=kontrakt.id))
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    return render_template('edit_kontrakt.html', kontrakt=kontrakt, evenementen=evenementen)

@app.route('/kontrakten/<int:id>/delete', methods=['POST'])
@beheerder_required
def delete_kontrakt(id):
    kontrakt = Kontrakt.query.get_or_404(id)
    
    # Check if kontrakt has related sponsoringen
    if kontrakt.sponsoringen:
        flash(f'Kontrakt "{kontrakt.kontrakt}" kan niet worden verwijderd omdat er nog sponsoringen aan gekoppeld zijn.', 'error')
        return redirect(url_for('kontrakten'))
    
    # Delete the kontrakt
    db.session.delete(kontrakt)
    db.session.commit()
    
    flash(f'Kontrakt "{kontrakt.kontrakt}" is succesvol verwijderd.', 'success')
    return redirect(url_for('kontrakten'))

@app.route('/kontrakten/export/excel')
@login_required
def export_kontrakten_excel():
    # Get filter parameters
    evenement_filter = request.args.get('evenement', '')
    
    # If no evenement filter is provided, default to the most recent event
    if not evenement_filter:
        latest_evenement = Evenement.query.order_by(Evenement.datum.desc()).first()
        if latest_evenement:
            evenement_filter = str(latest_evenement.id)
    
    query = Kontrakt.query
    if evenement_filter:
        query = query.filter(Kontrakt.evenement_id == evenement_filter)
    
    kontrakten = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrakten"
    
    # Headers
    headers = ['Evenement', 'Datum', 'Kontrakt', 'Bedrag', 'Aantal Sponsoringen']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, kontrakt in enumerate(kontrakten, 2):
        ws.cell(row=row, column=1, value=kontrakt.evenement.naam)
        ws.cell(row=row, column=2, value=kontrakt.evenement.datum.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=kontrakt.kontrakt)
        ws.cell(row=row, column=4, value=kontrakt.bedrag)
        ws.cell(row=row, column=5, value=len(kontrakt.sponsoringen))
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='kontrakten.xlsx')

@app.route('/kontrakten/export/pdf')
@login_required
def export_kontrakten_pdf():
    # Get filter parameters
    evenement_filter = request.args.get('evenement', '')
    
    # If no evenement filter is provided, default to the most recent event
    if not evenement_filter:
        latest_evenement = Evenement.query.order_by(Evenement.datum.desc()).first()
        if latest_evenement:
            evenement_filter = str(latest_evenement.id)
    
    query = Kontrakt.query
    if evenement_filter:
        query = query.filter(Kontrakt.evenement_id == evenement_filter)
    
    kontrakten = query.all()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30)
    story.append(Paragraph("Kontrakten Overzicht", title_style))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Evenement', 'Datum', 'Kontrakt', 'Bedrag', 'Sponsoringen']]
    for kontrakt in kontrakten:
        data.append([
            kontrakt.evenement.naam,
            kontrakt.evenement.datum.strftime('%d/%m/%Y'),
            kontrakt.kontrakt,
            f"€{format_european_currency(kontrakt.bedrag)}",
            str(len(kontrakt.sponsoringen))
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='kontrakten.pdf')

@app.route('/sponsors')
@login_required
def sponsors():
    # Get filter parameters
    bestuurslid_filter = request.args.get('bestuurslid', '')
    naam_filter = request.args.get('naam', '')
    kontaktpersoon_filter = request.args.get('kontaktpersoon', '')
    sort_key = request.args.get('sort', '')
    sort_dir = request.args.get('dir', 'asc')
    
    # Start with all sponsors
    query = Sponsor.query
    
    # Apply filters if provided
    if bestuurslid_filter:
        # Filter by bestuurslid name or initials
        query = query.join(Bestuurslid).filter(
            db.or_(
                Bestuurslid.naam.ilike(f'%{bestuurslid_filter}%'),
                Bestuurslid.initialen.ilike(f'%{bestuurslid_filter}%')
            )
        )
    if naam_filter:
        query = query.filter(Sponsor.naam.ilike(f'%{naam_filter}%'))
    if kontaktpersoon_filter:
        query = query.filter(Sponsor.kontaktpersoon.ilike(f'%{kontaktpersoon_filter}%'))
    
    sponsors = query.all()
    
    # Apply sorting
    def sort_key_fn(s):
        if sort_key == 'naam':
            return (s.naam or '').lower()
        if sort_key == 'kontaktpersoon':
            return (s.kontaktpersoon or '').lower()
        if sort_key == 'telefoon':
            return (s.telefoon or '').lower()
        if sort_key == 'email':
            return (s.email or '').lower()
        if sort_key == 'bestuurslid':
            if s.bestuurslid:
                return (s.bestuurslid.naam or s.bestuurslid.initialen or '').lower()
            return ''
        if sort_key == 'sponsoringen':
            return len(s.sponsoringen)
        # default: alphabetical by name
        return (s.naam or '').lower()
    
    if sort_key:
        sponsors = sorted(sponsors, key=sort_key_fn, reverse=(sort_dir == 'desc'))
    else:
        sponsors = sorted(sponsors, key=lambda s: (s.naam or '').lower())
    
    bestuursleden = Bestuurslid.query.all()
    
    # Sort bestuursleden by naam if available, otherwise by initialen
    bestuursleden = sorted(bestuursleden, key=lambda b: b.naam if b.naam else b.initialen)
    
    # Get all sponsors for dropdown options
    all_sponsors = Sponsor.query.all()
    
    return render_template('sponsors.html', 
                         sponsors=sponsors, 
                         all_sponsors=all_sponsors,
                         bestuursleden=bestuursleden,
                         selected_bestuurslid=bestuurslid_filter,
                         selected_naam=naam_filter,
                         selected_kontaktpersoon=kontaktpersoon_filter,
                         selected_sort=sort_key,
                         selected_dir=sort_dir)

@app.route('/sponsors/add', methods=['GET', 'POST'])
@gebruiker_required
def add_sponsor():
    if request.method == 'POST':
        sponsor = Sponsor(
            naam=request.form['naam'],
            straat=request.form['straat'],
            huisnummer=request.form['huisnummer'],
            postcode=request.form['postcode'],
            gemeente=request.form['gemeente'],
            kontaktpersoon=request.form['kontaktpersoon'],
            telefoon=request.form['telefoon'],
            email=request.form['email'],
            btw_nummer=request.form['btw_nummer'],
            opmerkingen=request.form.get('opmerkingen') or None,
            bestuurslid_id=request.form['bestuurslid_id'] if request.form['bestuurslid_id'] else None
        )
        db.session.add(sponsor)
        db.session.commit()
        flash('Sponsor succesvol toegevoegd!', 'success')
        return redirect(url_for('sponsors'))
    bestuursleden = Bestuurslid.query.all()
    return render_template('add_sponsor.html', bestuursleden=bestuursleden)

@app.route('/sponsors/<int:id>')
@login_required
def sponsor_detail(id):
    sponsor = Sponsor.query.get_or_404(id)
    
    # Calculate total bedrag
    total_bedrag = sum(s.facturatiebedrag_incl_btw for s in sponsor.sponsoringen if s.facturatiebedrag_incl_btw is not None)
    
    return render_template('sponsor_detail.html', sponsor=sponsor, total_bedrag=total_bedrag)

@app.route('/sponsors/<int:id>/edit', methods=['GET', 'POST'])
@gebruiker_required
def edit_sponsor(id):
    sponsor = Sponsor.query.get_or_404(id)
    if request.method == 'POST':
        sponsor.naam = request.form['naam']
        sponsor.straat = request.form['straat']
        sponsor.huisnummer = request.form['huisnummer']
        sponsor.postcode = request.form['postcode']
        sponsor.gemeente = request.form['gemeente']
        sponsor.kontaktpersoon = request.form['kontaktpersoon']
        sponsor.telefoon = request.form['telefoon']
        sponsor.email = request.form['email']
        sponsor.btw_nummer = request.form['btw_nummer']
        sponsor.opmerkingen = request.form.get('opmerkingen') or None
        sponsor.bestuurslid_id = request.form['bestuurslid_id'] if request.form['bestuurslid_id'] else None
        db.session.commit()
        flash('Sponsor succesvol bijgewerkt!', 'success')
        return redirect(url_for('sponsor_detail', id=sponsor.id))
    bestuursleden = Bestuurslid.query.all()
    return render_template('edit_sponsor.html', sponsor=sponsor, bestuursleden=bestuursleden)

@app.route('/sponsors/<int:id>/delete', methods=['POST'])
@beheerder_required
def delete_sponsor(id):
    sponsor = Sponsor.query.get_or_404(id)
    
    # Check if sponsor has related sponsoringen
    if sponsor.sponsoringen:
        flash(f'Sponsor "{sponsor.naam}" kan niet worden verwijderd omdat er nog sponsoringen aan gekoppeld zijn.', 'error')
        return redirect(url_for('sponsors'))
    
    # Delete the sponsor
    db.session.delete(sponsor)
    db.session.commit()
    
    flash(f'Sponsor "{sponsor.naam}" is succesvol verwijderd.', 'success')
    return redirect(url_for('sponsors'))

@app.route('/sponsors/export/excel')
@login_required
def export_sponsors_excel():
    # Get filter parameters
    bestuurslid_filter = request.args.get('bestuurslid', '')
    naam_filter = request.args.get('naam', '')
    kontaktpersoon_filter = request.args.get('kontaktpersoon', '')
    
    query = Sponsor.query
    if bestuurslid_filter:
        query = query.filter(Sponsor.bestuurslid_id == bestuurslid_filter)
    if naam_filter:
        query = query.filter(Sponsor.naam.contains(naam_filter))
    if kontaktpersoon_filter:
        query = query.filter(Sponsor.kontaktpersoon.contains(kontaktpersoon_filter))
    
    sponsors = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sponsors"
    
    # Headers
    headers = ['Naam', 'Kontaktpersoon', 'Telefoon', 'Email', 'Bestuurslid', 'Aantal Sponsoringen']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, sponsor in enumerate(sponsors, 2):
        ws.cell(row=row, column=1, value=sponsor.naam)
        ws.cell(row=row, column=2, value=sponsor.kontaktpersoon or '')
        ws.cell(row=row, column=3, value=sponsor.telefoon or '')
        ws.cell(row=row, column=4, value=sponsor.email or '')
        ws.cell(row=row, column=5, value=sponsor.bestuurslid.naam or sponsor.bestuurslid.initialen if sponsor.bestuurslid else '')
        ws.cell(row=row, column=6, value=len(sponsor.sponsoringen))
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='sponsors.xlsx')

@app.route('/sponsors/export/pdf')
@login_required
def export_sponsors_pdf():
    # Get filter parameters
    bestuurslid_filter = request.args.get('bestuurslid', '')
    naam_filter = request.args.get('naam', '')
    kontaktpersoon_filter = request.args.get('kontaktpersoon', '')
    
    query = Sponsor.query
    if bestuurslid_filter:
        query = query.filter(Sponsor.bestuurslid_id == bestuurslid_filter)
    if naam_filter:
        query = query.filter(Sponsor.naam.contains(naam_filter))
    if kontaktpersoon_filter:
        query = query.filter(Sponsor.kontaktpersoon.contains(kontaktpersoon_filter))
    
    sponsors = query.all()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30)
    story.append(Paragraph("Sponsors Overzicht", title_style))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Naam', 'Kontaktpersoon', 'Telefoon', 'Email', 'Bestuurslid', 'Sponsoringen']]
    for sponsor in sponsors:
        data.append([
            sponsor.naam,
            sponsor.kontaktpersoon or '',
            sponsor.telefoon or '',
            sponsor.email or '',
            sponsor.bestuurslid.naam or sponsor.bestuurslid.initialen if sponsor.bestuurslid else '',
            str(len(sponsor.sponsoringen))
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='sponsors.pdf')

@app.route('/sponsoringen')
@login_required
def sponsoringen():
    # Get filter parameters
    evenement_filter = request.args.get('evenement', '')
    kontrakt_filter = request.args.get('kontrakt', '')
    sponsor_filter = request.args.get('sponsor', '')
    logo_bezorgd_filter = request.args.get('logo_bezorgd', '')
    logo_afgewerkt_filter = request.args.get('logo_afgewerkt', '')
    gefactureerd_filter = request.args.get('gefactureerd', '')
    betaald_filter = request.args.get('betaald', '')
    sort_key = request.args.get('sort', '')  # '', 'evenement', 'kontrakt', 'sponsor', 'aangebracht', 'bedrag'
    sort_dir = request.args.get('dir', 'asc')  # 'asc' or 'desc'
    
    # If no evenement filter is provided, default to the most recent event
    if not evenement_filter:
        latest_evenement = Evenement.query.order_by(Evenement.datum.desc()).first()
        if latest_evenement:
            evenement_filter = str(latest_evenement.id)
    
    # Build query
    query = Sponsoring.query
    
    # Apply filters
    if evenement_filter:
        query = query.filter(Sponsoring.evenement_id == int(evenement_filter))
    if kontrakt_filter:
        query = query.filter(Sponsoring.kontrakt_id == int(kontrakt_filter))
    if sponsor_filter:
        # Check if sponsor_filter is a numeric ID or a name
        try:
            sponsor_id = int(sponsor_filter)
            # It's an ID, filter by sponsor ID
            query = query.filter(Sponsoring.sponsor_id == sponsor_id)
        except ValueError:
            # It's not a number, treat as name filter (partial match)
            query = query.join(Sponsor).filter(Sponsor.naam.ilike(f'%{sponsor_filter}%'))
    if logo_bezorgd_filter:
        if logo_bezorgd_filter == 'ja':
            query = query.filter(Sponsoring.logo_bezorgd == True)
        elif logo_bezorgd_filter == 'nee':
            query = query.filter(Sponsoring.logo_bezorgd == False)
    if logo_afgewerkt_filter:
        if logo_afgewerkt_filter == 'ja':
            query = query.filter(Sponsoring.logo_afgewerkt == True)
        elif logo_afgewerkt_filter == 'nee':
            query = query.filter(Sponsoring.logo_afgewerkt == False)
    if gefactureerd_filter:
        if gefactureerd_filter == 'ja':
            query = query.filter(Sponsoring.gefactureerd == True)
        elif gefactureerd_filter == 'nee':
            query = query.filter(Sponsoring.gefactureerd == False)
    if betaald_filter:
        if betaald_filter == 'ja':
            query = query.filter(Sponsoring.betaald == True)
        elif betaald_filter == 'nee':
            query = query.filter(Sponsoring.betaald == False)
    
    sponsoringen = query.all()

    # Apply sorting in Python to support derived fields like display amount
    def safe_lower(text):
        return (text or '').lower()

    def sort_key_fn(s):
        if sort_key == 'evenement':
            return safe_lower(getattr(s.evenement, 'naam', ''))
        if sort_key == 'kontrakt':
            return safe_lower(getattr(s.kontrakt, 'kontrakt', ''))
        if sort_key == 'sponsor':
            return safe_lower(getattr(s.sponsor, 'naam', ''))
        if sort_key == 'aangebracht':
            display = s.aangebracht_door.naam or s.aangebracht_door.initialen if s.aangebracht_door else ''
            return safe_lower(display)
        if sort_key == 'bedrag':
            return float(get_display_amount(s) or 0)
        if sort_key == 'bedrag_incl':
            return float(s.facturatiebedrag_incl_btw or 0)
        # default fallback: newest first like before
        return s.id

    if sort_key:
        sponsoringen = sorted(sponsoringen, key=sort_key_fn, reverse=(sort_dir == 'desc'))
    else:
        sponsoringen = sorted(sponsoringen, key=lambda s: s.id, reverse=True)
    
    # Calculate totals with fallback logic (same as template logic)
    total_netto_bedrag = sum(get_display_amount(s) for s in sponsoringen)
    
    total_facturatie_bedrag = sum(s.facturatiebedrag_incl_btw for s in sponsoringen if s.facturatiebedrag_incl_btw is not None)
    
    # Get filter options
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    
    # Get kontrakten based on evenement filter
    if evenement_filter:
        kontrakten = Kontrakt.query.filter_by(evenement_id=int(evenement_filter)).all()
    else:
        # Show all kontrakten when no evenement filter
        kontrakten = Kontrakt.query.join(Evenement).order_by(Evenement.datum.desc(), Kontrakt.kontrakt).all()
    
    sponsors = Sponsor.query.all()
    
    return render_template('sponsoringen.html',
                         sponsoringen=sponsoringen,
                         evenementen=evenementen,
                         kontrakten=kontrakten,
                         sponsors=sponsors,
                         total_netto_bedrag=total_netto_bedrag,
                         total_facturatie_bedrag=total_facturatie_bedrag,
                         selected_evenement=evenement_filter,
                         selected_kontrakt=kontrakt_filter,
                         selected_sponsor=sponsor_filter,
                         selected_logo_bezorgd=logo_bezorgd_filter,
                         selected_logo_afgewerkt=logo_afgewerkt_filter,
                         selected_gefactureerd=gefactureerd_filter,
                         selected_betaald=betaald_filter,
                         selected_sort=sort_key,
                         selected_dir=sort_dir)

@app.route('/sponsoringen/add', methods=['GET', 'POST'])
@gebruiker_required
def add_sponsoring():
    if request.method == 'POST':
        # Handle file uploads
        logo_origineel = None
        logo_afgewerkt = None
        
        if 'logo_origineel' in request.files:
            file = request.files['logo_origineel']
            if file and file.filename:
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4()}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                logo_origineel = unique_filename
        
        if 'logo_afgewerkt' in request.files:
            file = request.files['logo_afgewerkt']
            if file and file.filename:
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4()}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                logo_afgewerkt = unique_filename
        
        sponsoring = Sponsoring(
            evenement_id=request.form['evenement_id'],
            kontrakt_id=request.form['kontrakt_id'],
            sponsor_id=request.form['sponsor_id'],
            aangebracht_door_id=request.form['aangebracht_door_id'],
            bedrag_kaarten=float(request.form['bedrag_kaarten']) if request.form['bedrag_kaarten'] else None,
            netto_bedrag_excl_btw=float(request.form['netto_bedrag_excl_btw']) if request.form['netto_bedrag_excl_btw'] else None,
            facturatiebedrag_incl_btw=float(request.form['facturatiebedrag_incl_btw']) if request.form['facturatiebedrag_incl_btw'] else None,
            gefactureerd=bool(request.form.get('gefactureerd')),
            betaald=bool(request.form.get('betaald')),
            opmerkingen=request.form['opmerkingen'],
            logo_bezorgd=bool(logo_origineel),
            logo_afgewerkt=bool(logo_afgewerkt),
            logo_origineel=logo_origineel,
            logo_afgewerkt_file=logo_afgewerkt
        )
        db.session.add(sponsoring)
        db.session.commit()
        flash('Sponsoring succesvol toegevoegd!', 'success')
        return redirect(url_for('sponsoringen'))
    
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    kontrakten = Kontrakt.query.all()
    sponsors = Sponsor.query.all()
    bestuursleden = Bestuurslid.query.all()
    
    # Sort bestuursleden by naam if available, otherwise by initialen
    bestuursleden = sorted(bestuursleden, key=lambda b: b.naam if b.naam else b.initialen)
    
    return render_template('add_sponsoring.html', 
                         evenementen=evenementen,
                         kontrakten=kontrakten,
                         sponsors=sponsors,
                         bestuursleden=bestuursleden)

@app.route('/sponsoringen/<int:id>')
@login_required
def sponsoring_detail(id):
    sponsoring = Sponsoring.query.get_or_404(id)
    
    # Get referrer information to preserve filters
    referrer = request.referrer
    back_url = url_for('sponsoringen')
    
    # If coming from sponsoringen page, try to preserve filters
    if referrer and 'sponsoringen' in referrer:
        # Extract query parameters from referrer
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(referrer)
        query_params = parse_qs(parsed.query)
        
        # Build back URL with preserved filters
        filter_params = {}
        for key in ['evenement', 'kontrakt', 'sponsor', 'logo_bezorgd', 'logo_afgewerkt', 'gefactureerd', 'betaald']:
            if key in query_params:
                filter_params[key] = query_params[key][0]
        
        if filter_params:
            back_url = url_for('sponsoringen', **filter_params)
    
    return render_template('sponsoring_detail.html', sponsoring=sponsoring, back_url=back_url)

@app.route('/sponsoringen/<int:id>/edit', methods=['GET', 'POST'])
@gebruiker_required
def edit_sponsoring(id):
    sponsoring = Sponsoring.query.get_or_404(id)
    if request.method == 'POST':
        # Handle file uploads
        logo_origineel = sponsoring.logo_origineel
        logo_afgewerkt = sponsoring.logo_afgewerkt_file
        
        if 'logo_origineel' in request.files:
            file = request.files['logo_origineel']
            if file and file.filename:
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4()}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                logo_origineel = unique_filename
        
        if 'logo_afgewerkt' in request.files:
            file = request.files['logo_afgewerkt']
            if file and file.filename:
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4()}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                logo_afgewerkt = unique_filename
        
        sponsoring.kontrakt_id = request.form['kontrakt_id']
        sponsoring.sponsor_id = request.form['sponsor_id']
        sponsoring.aangebracht_door_id = request.form['aangebracht_door_id']
        sponsoring.bedrag_kaarten = float(request.form['bedrag_kaarten']) if request.form['bedrag_kaarten'] else None
        sponsoring.netto_bedrag_excl_btw = float(request.form['netto_bedrag_excl_btw']) if request.form['netto_bedrag_excl_btw'] else None
        sponsoring.facturatiebedrag_incl_btw = float(request.form['facturatiebedrag_incl_btw']) if request.form['facturatiebedrag_incl_btw'] else None
        sponsoring.gefactureerd = bool(request.form.get('gefactureerd'))
        sponsoring.betaald = bool(request.form.get('betaald'))
        sponsoring.opmerkingen = request.form['opmerkingen']
        sponsoring.logo_bezorgd = bool(logo_origineel)
        sponsoring.logo_afgewerkt = bool(logo_afgewerkt)
        sponsoring.logo_origineel = logo_origineel
        sponsoring.logo_afgewerkt_file = logo_afgewerkt
        
        db.session.commit()
        flash('Sponsoring succesvol bijgewerkt!', 'success')
        return redirect(url_for('sponsoring_detail', id=sponsoring.id))
    
    evenementen = Evenement.query.order_by(Evenement.datum.desc()).all()
    kontrakten = Kontrakt.query.filter_by(evenement_id=sponsoring.evenement_id).all()
    sponsors = Sponsor.query.all()
    bestuursleden = Bestuurslid.query.all()
    
    # Sort bestuursleden by naam if available, otherwise by initialen
    bestuursleden = sorted(bestuursleden, key=lambda b: b.naam if b.naam else b.initialen)
    
    return render_template('edit_sponsoring.html', 
                         sponsoring=sponsoring,
                         evenementen=evenementen,
                         kontrakten=kontrakten,
                         sponsors=sponsors,
                         bestuursleden=bestuursleden)

@app.route('/sponsoringen/<int:id>/delete', methods=['POST'])
@beheerder_required
def delete_sponsoring(id):
    sponsoring = Sponsoring.query.get_or_404(id)
    
    # Delete logo files if they exist
    if sponsoring.logo_origineel:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], sponsoring.logo_origineel))
        except:
            pass
    if sponsoring.logo_afgewerkt_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], sponsoring.logo_afgewerkt_file))
        except:
            pass
    
    # Delete the sponsoring
    db.session.delete(sponsoring)
    db.session.commit()
    
    flash(f'Sponsoring is succesvol verwijderd.', 'success')
    return redirect(url_for('sponsoringen'))

@app.route('/sponsoringen/export/excel')
@login_required
def export_sponsoringen_excel():
    # Get filter parameters
    evenement_filter = request.args.get('evenement', '')
    kontrakt_filter = request.args.get('kontrakt', '')
    sponsor_filter = request.args.get('sponsor', '')
    logo_bezorgd_filter = request.args.get('logo_bezorgd', '')
    logo_afgewerkt_filter = request.args.get('logo_afgewerkt', '')
    gefactureerd_filter = request.args.get('gefactureerd', '')
    betaald_filter = request.args.get('betaald', '')
    
    # If no evenement filter is provided, default to the most recent event
    if not evenement_filter:
        latest_evenement = Evenement.query.order_by(Evenement.datum.desc()).first()
        if latest_evenement:
            evenement_filter = str(latest_evenement.id)
    
    query = Sponsoring.query
    if evenement_filter:
        query = query.filter(Sponsoring.evenement_id == evenement_filter)
    if kontrakt_filter:
        query = query.filter(Sponsoring.kontrakt_id == kontrakt_filter)
    if sponsor_filter:
        # Check if sponsor_filter is a numeric ID or a name
        try:
            sponsor_id = int(sponsor_filter)
            # It's an ID, filter by sponsor ID
            query = query.filter(Sponsoring.sponsor_id == sponsor_id)
        except ValueError:
            # It's not a number, treat as name filter (partial match)
            query = query.join(Sponsor).filter(Sponsor.naam.ilike(f'%{sponsor_filter}%'))
    if logo_bezorgd_filter:
        if logo_bezorgd_filter == 'ja':
            query = query.filter(Sponsoring.logo_origineel.isnot(None))
        elif logo_bezorgd_filter == 'nee':
            query = query.filter(Sponsoring.logo_origineel.is_(None))
    if logo_afgewerkt_filter:
        if logo_afgewerkt_filter == 'ja':
            query = query.filter(Sponsoring.logo_afgewerkt_file.isnot(None))
        elif logo_afgewerkt_filter == 'nee':
            query = query.filter(Sponsoring.logo_afgewerkt_file.is_(None))
    if gefactureerd_filter:
        if gefactureerd_filter == 'ja':
            query = query.filter(Sponsoring.gefactureerd == True)
        elif gefactureerd_filter == 'nee':
            query = query.filter(Sponsoring.gefactureerd == False)
    if betaald_filter:
        if betaald_filter == 'ja':
            query = query.filter(Sponsoring.betaald == True)
        elif betaald_filter == 'nee':
            query = query.filter(Sponsoring.betaald == False)
    
    sponsoringen = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sponsoringen"
    
    # Headers
    headers = ['Evenement', 'Datum', 'Kontrakt', 'Sponsor', 'Aangebracht door', 'Bedrag kaarten', 'Netto bedrag excl BTW', 'Facturatiebedrag incl BTW', 'Gefactureerd', 'Betaald']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, sponsoring in enumerate(sponsoringen, 2):
        ws.cell(row=row, column=1, value=sponsoring.evenement.naam)
        ws.cell(row=row, column=2, value=sponsoring.evenement.datum.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=sponsoring.kontrakt.kontrakt)
        ws.cell(row=row, column=4, value=sponsoring.sponsor.naam)
        ws.cell(row=row, column=5, value=sponsoring.aangebracht_door.naam or sponsoring.aangebracht_door.initialen)
        ws.cell(row=row, column=6, value=sponsoring.bedrag_kaarten or 0)
        ws.cell(row=row, column=7, value=sponsoring.netto_bedrag_excl_btw or 0)
        ws.cell(row=row, column=8, value=sponsoring.facturatiebedrag_incl_btw or 0)
        ws.cell(row=row, column=9, value='Ja' if sponsoring.gefactureerd else 'Nee')
        ws.cell(row=row, column=10, value='Ja' if sponsoring.betaald else 'Nee')
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='sponsoringen.xlsx')

@app.route('/sponsoringen/export/pdf')
@login_required
def export_sponsoringen_pdf():
    # Get filter parameters
    evenement_filter = request.args.get('evenement', '')
    kontrakt_filter = request.args.get('kontrakt', '')
    sponsor_filter = request.args.get('sponsor', '')
    logo_bezorgd_filter = request.args.get('logo_bezorgd', '')
    logo_afgewerkt_filter = request.args.get('logo_afgewerkt', '')
    gefactureerd_filter = request.args.get('gefactureerd', '')
    betaald_filter = request.args.get('betaald', '')
    
    # If no evenement filter is provided, default to the most recent event
    if not evenement_filter:
        latest_evenement = Evenement.query.order_by(Evenement.datum.desc()).first()
        if latest_evenement:
            evenement_filter = str(latest_evenement.id)
    
    query = Sponsoring.query
    if evenement_filter:
        query = query.filter(Sponsoring.evenement_id == evenement_filter)
    if kontrakt_filter:
        query = query.filter(Sponsoring.kontrakt_id == kontrakt_filter)
    if sponsor_filter:
        # Check if sponsor_filter is a numeric ID or a name
        try:
            sponsor_id = int(sponsor_filter)
            # It's an ID, filter by sponsor ID
            query = query.filter(Sponsoring.sponsor_id == sponsor_id)
        except ValueError:
            # It's not a number, treat as name filter (partial match)
            query = query.join(Sponsor).filter(Sponsor.naam.ilike(f'%{sponsor_filter}%'))
    if logo_bezorgd_filter:
        if logo_bezorgd_filter == 'ja':
            query = query.filter(Sponsoring.logo_origineel.isnot(None))
        elif logo_bezorgd_filter == 'nee':
            query = query.filter(Sponsoring.logo_origineel.is_(None))
    if logo_afgewerkt_filter:
        if logo_afgewerkt_filter == 'ja':
            query = query.filter(Sponsoring.logo_afgewerkt_file.isnot(None))
        elif logo_afgewerkt_filter == 'nee':
            query = query.filter(Sponsoring.logo_afgewerkt_file.is_(None))
    if gefactureerd_filter:
        if gefactureerd_filter == 'ja':
            query = query.filter(Sponsoring.gefactureerd == True)
        elif gefactureerd_filter == 'nee':
            query = query.filter(Sponsoring.gefactureerd == False)
    if betaald_filter:
        if betaald_filter == 'ja':
            query = query.filter(Sponsoring.betaald == True)
        elif betaald_filter == 'nee':
            query = query.filter(Sponsoring.betaald == False)
    
    sponsoringen = query.all()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30)
    story.append(Paragraph("Sponsoringen Overzicht", title_style))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Evenement', 'Datum', 'Kontrakt', 'Sponsor', 'Aangebracht door', 'Bedrag kaarten', 'Netto bedrag excl BTW', 'Facturatiebedrag incl BTW', 'Gefactureerd', 'Betaald']]
    for sponsoring in sponsoringen:
        data.append([
            sponsoring.evenement.naam,
            sponsoring.evenement.datum.strftime('%d/%m/%Y'),
            sponsoring.kontrakt.kontrakt,
            sponsoring.sponsor.naam,
            sponsoring.aangebracht_door.naam or sponsoring.aangebracht_door.initialen,
            f"€{format_european_currency(sponsoring.bedrag_kaarten or 0)}",
            f"€{format_european_currency(sponsoring.netto_bedrag_excl_btw or 0)}",
            f"€{format_european_currency(sponsoring.facturatiebedrag_incl_btw or 0)}",
            'Ja' if sponsoring.gefactureerd else 'Nee',
            'Ja' if sponsoring.betaald else 'Nee'
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='sponsoringen.pdf')

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# API endpoint for dynamic kontrakt filtering
@app.route('/api/kontrakten/<int:evenement_id>')
@login_required
def get_kontrakten(evenement_id):
    kontrakten = Kontrakt.query.filter_by(evenement_id=evenement_id).all()
    return jsonify([{'id': k.id, 'kontrakt': k.kontrakt} for k in kontrakten])

@app.route('/download-logos')
@login_required
def download_logos():
    # Get all sponsoringen with afgewerkte logos
    sponsoringen = Sponsoring.query.filter(Sponsoring.logo_afgewerkt_file.isnot(None)).all()
    
    if not sponsoringen:
        flash('Geen afgewerkte logo\'s gevonden om te downloaden.', 'warning')
        return redirect(url_for('sponsoringen'))
    
    # Create temporary zip file
    temp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(temp_dir, 'afgewerkte_logos.zip')
    
    try:
        files_added = 0
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for sponsoring in sponsoringen:
                if sponsoring.logo_afgewerkt_file:
                    # Get the file path
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], sponsoring.logo_afgewerkt_file)
                    
                    if os.path.exists(file_path):
                        # Create a descriptive filename
                        evenement_naam = sponsoring.evenement.naam.replace(' ', '_').replace('/', '_')
                        sponsor_naam = sponsoring.sponsor.naam.replace(' ', '_').replace('/', '_')
                        kontrakt_naam = sponsoring.kontrakt.kontrakt.replace(' ', '_').replace('/', '_')
                        
                        # Get file extension
                        _, ext = os.path.splitext(sponsoring.logo_afgewerkt_file)
                        
                        # Create descriptive filename
                        filename = f"{evenement_naam}_{sponsor_naam}_{kontrakt_naam}_afgewerkt{ext}"
                        
                        # Add file to zip
                        zipf.write(file_path, filename)
                        files_added += 1
        
        if files_added == 0:
            flash('Geen afgewerkte logo bestanden gevonden om te downloaden.', 'warning')
            return redirect(url_for('sponsoringen'))
        
        # Send the zip file
        return send_file(
            zip_path,
            as_attachment=True,
            download_name=f'afgewerkte_logos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip',
            mimetype='application/zip'
        )
        
    except Exception as e:
        flash(f'Fout bij het maken van het zipbestand: {str(e)}', 'error')
        return redirect(url_for('sponsoringen'))
    finally:
        # Clean up temporary directory
        try:
            if os.path.exists(zip_path):
                os.remove(zip_path)
            os.rmdir(temp_dir)
        except:
            pass


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Ensure 'opmerkingen' column exists on Sponsor for existing databases
        try:
            from sqlalchemy import text
            result = db.session.execute(text("PRAGMA table_info(sponsor)"))
            cols = [row[1] for row in result]
            if 'opmerkingen' not in cols:
                db.session.execute(text("ALTER TABLE sponsor ADD COLUMN opmerkingen TEXT"))
                db.session.commit()
        except Exception:
            db.session.rollback()

        # Relax NOT NULL constraint on bestuurslid.naam if still present (SQLite)
        try:
            info = db.session.execute(text("PRAGMA table_info(bestuurslid)")).fetchall()
            naam_info = next((c for c in info if c[1] == 'naam'), None)
            # PRAGMA table_info columns: cid, name, type, notnull, dflt_value, pk
            if naam_info is not None and int(naam_info[3]) == 1:
                # Rebuild table with naam nullable
                db.session.execute(text("PRAGMA foreign_keys=OFF"))
                db.session.execute(text("BEGIN TRANSACTION"))
                db.session.execute(text(
                    "CREATE TABLE IF NOT EXISTS bestuurslid_new (\n"
                    "    id INTEGER PRIMARY KEY,\n"
                    "    initialen VARCHAR(10) NOT NULL,\n"
                    "    naam VARCHAR(100) NULL\n"
                    ")"
                ))
                db.session.execute(text(
                    "INSERT INTO bestuurslid_new (id, initialen, naam) "
                    "SELECT id, initialen, naam FROM bestuurslid"
                ))
                db.session.execute(text("DROP TABLE bestuurslid"))
                db.session.execute(text("ALTER TABLE bestuurslid_new RENAME TO bestuurslid"))
                db.session.execute(text("COMMIT"))
                db.session.execute(text("PRAGMA foreign_keys=ON"))
                db.session.commit()
        except Exception:
            db.session.rollback()

@app.route('/sponsors/add-ajax', methods=['POST'])
@gebruiker_required
def add_sponsor_ajax():
    try:
        sponsor = Sponsor(
            naam=request.form['naam'],
            kontaktpersoon=request.form.get('kontaktpersoon', ''),
            telefoon=request.form.get('telefoon', ''),
            email=request.form.get('email', ''),
            bestuurslid_id=request.form.get('bestuurslid_id') if request.form.get('bestuurslid_id') else None
        )
        db.session.add(sponsor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'sponsor_id': sponsor.id,
            'sponsor_name': sponsor.naam
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/bestuursleden/add-ajax', methods=['POST'])
@gebruiker_required
def add_bestuurslid_ajax():
    try:
        bestuurslid = Bestuurslid(
            initialen=request.form['initialen'],
            naam=request.form.get('naam', '') if request.form.get('naam') else None
        )
        db.session.add(bestuurslid)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'bestuurslid_id': bestuurslid.id,
            'bestuurslid_name': bestuurslid.naam or bestuurslid.initialen
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })
        
        # Create default admin user if no users exist
        if Gebruiker.query.count() == 0:
            admin = Gebruiker(
                email='admin@kampanje.be',
                rol='beheerder'
            )
            admin.set_password('admin123')  # Default password
            db.session.add(admin)
            db.session.commit()
            print("Default admin user created: admin@kampanje.be / admin123")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5100)
